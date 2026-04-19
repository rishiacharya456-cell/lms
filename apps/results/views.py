from apps.results.models import Result
from apps.results.utils import calculate_grade
from apps.exams.models import ExamRegistration


def upload_marks(request):

    school = request.user.assigned_school
    selected_class = request.GET.get('class')

    regs = ExamRegistration.objects.filter(
        school=school,
        status='approved',
        is_verified=True
    ).select_related('student', 'course')

    if selected_class:
        regs = regs.filter(student__student_class=selected_class)

    if request.method == "POST":

        for reg in regs:

            th = int(request.POST.get(f"th_{reg.id}", 0))
            internal = int(request.POST.get(f"in_{reg.id}", 0))

            total = th + internal
            grade, gp = calculate_grade(total)

            Result.objects.update_or_create(
                student=reg.student,
                course=reg.course,
                defaults={
                    "school": school,
                    "theory_marks": th,
                    "internal_marks": internal,
                    "total": total,
                    "grade": grade,
                    "grade_point": gp,
                    "is_pass": total >= reg.course.pass_marks,
                    "uploaded_by": request.user
                }
            )

    return render(request, 'results/upload_marks.html', {
        "regs": regs
    })
    
    
    
def publish_results(request):

    school_id = request.POST.get('school')

    Result.objects.filter(
        school_id=school_id
    ).update(is_locked=True)

    return redirect('admin_results')



def student_result(request):

    student = request.user

    regs = ExamRegistration.objects.filter(
        student=student,
        is_verified=True
    )

    if not regs.exists():
        return render(request, 'results/student_result.html', {
            "detained": True
        })

    results = Result.objects.filter(student=student)

    total_gp = sum(r.grade_point for r in results)
    cgpa = total_gp / results.count() if results else 0

    return render(request, 'results/student_result.html', {
        "results": results,
        "cgpa": round(cgpa, 2)
    })
    
    
from django.shortcuts import render
from apps.schools.models import School
from apps.accounts.models import User
from apps.results.models import Result, ResultPublish
from apps.exams.models import ExamRegistration


def admin_results(request):

    schools = School.objects.all()

    # ✅ FETCH PUBLISH STATUS (OPTIMIZED)
    publish_map = {
        p.school_id: p.is_published
        for p in ResultPublish.objects.all()
    }

    for s in schools:

        # =========================
        # TOTAL STUDENTS
        # =========================
        students_qs = User.objects.filter(
            role='student',
            school=s
        )
        total_students = students_qs.count()

        # =========================
        # ELIGIBLE STUDENTS
        # =========================
        eligible_students = ExamRegistration.objects.filter(
            school=s,
            status='approved',
            is_verified=True
        ).values('student').distinct()

        eligible_count = eligible_students.count()

        # =========================
        # RESULTS COUNT (UNIQUE STUDENTS)
        # =========================
        results_count = Result.objects.filter(
            school=s
        ).values('student').distinct().count()

        # =========================
        # DETAINED (CORRECT LOGIC)
        # =========================
        detained_count = total_students - eligible_count

        # =========================
        # ATTACH DATA TO SCHOOL
        # =========================
        s.total_students = total_students
        s.total_results = results_count
        s.detained_count = detained_count

        # 🔥 NEW: PUBLISH STATUS
        s.is_published = publish_map.get(s.id, False)

    return render(request, 'results/admin_results.html', {
        "schools": schools
    })
    
    
    
from django.shortcuts import render, get_object_or_404
from apps.accounts.models import User
from apps.exams.models import ExamRegistration
from apps.results.models import Result
from apps.results.utils import calculate_grade
from apps.schools.models import School


def admin_school_results(request, school_id):

    school = get_object_or_404(School, id=school_id)

    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')

    # ✅ ALL STUDENTS (ALWAYS SHOW)
    students = User.objects.filter(
        role='student',
        school=school
    )

    if selected_class:
        students = students.filter(student_class=str(selected_class))

    if selected_section:
        students = students.filter(section=str(selected_section))

    # ✅ ALL REGISTRATIONS
    all_regs = ExamRegistration.objects.filter(
        school=school
    ).select_related('student', 'course')

    # ✅ TEMPLATE-FRIENDLY LOOKUP (IMPORTANT FIX)
    reg_lookup = {}
    for reg in all_regs:
        reg_lookup[str(reg.student_id)] = reg

    # ✅ ELIGIBLE STUDENTS (ONLY FOR SAVING MARKS)
    eligible_regs = [
        r for r in all_regs
        if r.status == "approved" and r.is_verified
    ]

    if selected_class:
        eligible_regs = [
            r for r in eligible_regs
            if str(r.student.student_class) == str(selected_class)
        ]

    if selected_section:
        eligible_regs = [
            r for r in eligible_regs
            if str(r.student.section) == str(selected_section)
        ]

    # ✅ SAVE MARKS (ONLY TH)
    if request.method == "POST":
        for reg in eligible_regs:

            th = int(request.POST.get(f"th_{reg.id}", 0))

            total = th
            grade, gp = calculate_grade(total)

            Result.objects.update_or_create(
                student=reg.student,
                course=reg.course,
                defaults={
                    "school": school,
                    "theory_marks": th,
                    "internal_marks": 0,
                    "total": total,
                    "grade": grade,
                    "grade_point": gp,
                    "is_pass": total >= reg.course.pass_marks
                }
            )

    # ✅ CLASS LIST
    classes = User.objects.filter(
        role='student',
        school=school
    ).values_list('student_class', flat=True).distinct()

    # ✅ SECTION LIST
    sections = students.values_list('section', flat=True).distinct()

    return render(request, 'results/admin_school_results.html', {
    "school": school,
    "students": students,
    "regs": all_regs,
    "verified_ids": verified_ids,   # ✅ NEW
    "result_map": result_map,       # ✅ NEW
    "classes": classes,
    "sections": sections,
    "selected_class": selected_class,
    "selected_section": selected_section
})

from django.shortcuts import redirect
from django.utils import timezone
from apps.results.models import Result, ResultPublish


def publish_results(request):

    if request.method == "POST":

        school_id = request.POST.get('school')

        # ✅ 1. LOCK ALL RESULTS
        Result.objects.filter(
            school_id=school_id
        ).update(is_locked=True)

        # ✅ 2. CREATE / UPDATE PUBLISH STATUS
        publish_obj, created = ResultPublish.objects.get_or_create(
            school_id=school_id
        )

        # 🚫 prevent duplicate publish
        if not publish_obj.is_published:
            publish_obj.is_published = True
            publish_obj.published_at = timezone.now()
            publish_obj.save()

    return redirect('admin_results')

from django.shortcuts import render
from apps.results.models import Result, ResultPublish
from apps.courses.models import CourseRegistration
from apps.results.utils import calculate_grade


def school_results(request):

    school = request.user.school
    selected_class = request.GET.get('class')

    # =========================
    # 🔒 CHECK PUBLISH STATUS
    # =========================
    is_published = ResultPublish.objects.filter(
        school=school,
        is_published=True
    ).exists()

    if not is_published:
        return render(request, "results/school_results.html", {
            "not_published": True
        })

    # =========================
    # RESULTS (THEORY)
    # =========================
    results = Result.objects.filter(
        school=school,
        is_locked=True
    ).select_related('student', 'course')

    if selected_class:
        results = results.filter(student__student_class=selected_class)

    # =========================
    # INTERNAL MARKS
    # =========================
    registrations = CourseRegistration.objects.filter(
        school=school,
        status='approved'
    ).select_related('student', 'course')

    reg_map = {
        (r.student_id, r.course_id): r
        for r in registrations
    }

    # =========================
    # FINAL DATA
    # =========================
    result_data = []

    for r in results:

        student = r.student
        course = r.course

        reg = reg_map.get((student.id, course.id))

        theory = r.theory_marks or 0
        internal = reg.internal_obtained if reg else 0

        obtained = theory + internal

        grade, gp = calculate_grade(obtained)

        is_pass = obtained >= course.pass_marks

        result_data.append({
            "student": student,
            "course": course,
            "full_marks": course.full_marks,
            "pass_marks": course.pass_marks,
            "theory": theory,
            "internal": internal,
            "obtained": obtained,
            "grade": grade,
            "gp": gp,
            "is_pass": is_pass
        })

    # =========================
    # CLASS FILTER LIST
    # =========================
    classes = (
        Result.objects
        .filter(school=school, is_locked=True)
        .values_list('student__student_class', flat=True)
        .distinct()
        .order_by('student__student_class')
    )

    return render(request, 'results/school_results.html', {
        "result_data": result_data,
        "classes": classes,
        "selected_class": selected_class,
        "not_published": False
    })
    

from django.shortcuts import render, get_object_or_404, redirect
from apps.accounts.models import User
from apps.exams.models import ExamRegistration
from apps.results.models import Result
from apps.results.utils import calculate_grade
from apps.schools.models import School


from django.shortcuts import render, get_object_or_404, redirect
from apps.accounts.models import User
from apps.exams.models import ExamRegistration
from apps.results.models import Result
from apps.results.utils import calculate_grade
from apps.schools.models import School

def admin_school_results(request, school_id):

    school = get_object_or_404(School, id=school_id)
    selected_class = request.GET.get('class')

    students = User.objects.filter(
        role='student',
        school=school
    )

    if not selected_class:
        first = students.first()
        selected_class = first.student_class if first else None

    students = students.filter(student_class=selected_class)

    regs = ExamRegistration.objects.filter(
        school=school,
        status='approved',
        is_verified=True,
        student__student_class=selected_class
    ).select_related('student', 'course')

    # ✅ SAVE MARKS (UNCHANGED BUT SAFE)
    if request.method == "POST":
        for reg in regs:
            th_value = request.POST.get(f"th_{reg.id}")

            if th_value is not None and th_value != "":
                th = int(th_value)

                total = th
                grade, gp = calculate_grade(total)

                Result.objects.update_or_create(
                    student=reg.student,
                    course=reg.course,
                    defaults={
                        "school": school,
                        "theory_marks": th,
                        "internal_marks": 0,
                        "total": total,
                        "grade": grade,
                        "grade_point": gp,
                        "is_pass": total >= reg.course.pass_marks
                    }
                )

        return redirect(request.path)

    # ✅ LOAD RESULTS
    results = Result.objects.filter(
        school=school,
        student__student_class=selected_class
    )

    # 🔥 FIX: correct mapping
    result_map = {
        (r.student_id, r.course_id): r
        for r in results
    }

    # 🔥 CLEAN DATA (IMPORTANT)
    student_data = []

    for student in students:

        student_regs = regs.filter(student=student)

        if student_regs.exists():
            for reg in student_regs:

                result = result_map.get((student.id, reg.course.id))

                student_data.append({
                    "student": student,
                    "course": reg.course,
                    "reg_id": reg.id,
                    "is_verified": True,
                    "marks": result.theory_marks if result else ""
                })

        else:
            # ❌ Not paid → detained
            student_data.append({
                "student": student,
                "course": None,
                "reg_id": None,
                "is_verified": False,
                "marks": ""
            })

    classes = User.objects.filter(
        role='student',
        school=school
    ).values_list('student_class', flat=True).distinct()

    return render(request, 'results/admin_school_results.html', {
        "school": school,
        "student_data": student_data,  # 🔥 USE THIS
        "classes": classes,
        "selected_class": selected_class,
    })
    



from django.shortcuts import render, redirect
from apps.accounts.models import User
from apps.exams.models import ExamRegistration
from apps.results.models import Result
from apps.results.utils import calculate_grade


def trainee_results(request):

    trainee = request.user

    # ✅ FIX: trainer uses assigned_school
    school = trainee.assigned_school

    # ✅ SAFETY CHECK
    if not school:
        return render(request, "results/trainee_results.html", {
            "student_data": [],
            "classes": [],
            "sections": [],
            "selected_class": None,
            "selected_section": None,
            "error": "No school assigned to this trainer"
        })

    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')

    # ✅ BASE STUDENTS (CLEAN DATA)
    base_students = User.objects.filter(
        role='student',
        school=school
    ).exclude(
        student_class__isnull=True
    ).exclude(
        student_class__exact=""
    ).exclude(
        section__isnull=True
    ).exclude(
        section__exact=""
    )

    # ✅ FILTER STUDENTS
    students = base_students

    if selected_class:
        students = students.filter(student_class=selected_class)

    if selected_section:
        students = students.filter(section=selected_section)

    # ✅ ALL REGISTRATIONS
    regs = ExamRegistration.objects.filter(
        school=school
    ).select_related('student', 'course')

    # ✅ ELIGIBLE (paid + verified)
    eligible_regs = regs.filter(
        status='approved',
        is_verified=True
    )

    if selected_class:
        eligible_regs = eligible_regs.filter(student__student_class=selected_class)

    if selected_section:
        eligible_regs = eligible_regs.filter(student__section=selected_section)

    # ✅ RESULTS
    results = Result.objects.filter(school=school)

    result_map = {
        (r.student_id, r.course_id): r
        for r in results
    }

    # ✅ SAVE MARKS (SAFE)
    if request.method == "POST":

        for reg in eligible_regs:

            value = request.POST.get(f"th_{reg.id}")

            if not value:
                continue

            # 🚫 DO NOT OVERWRITE EXISTING
            if (reg.student.id, reg.course.id) in result_map:
                continue

            th = int(value)
            total = th
            grade, gp = calculate_grade(total)

            Result.objects.create(
                student=reg.student,
                course=reg.course,
                school=school,
                theory_marks=th,
                internal_marks=0,
                total=total,
                grade=grade,
                grade_point=gp,
                is_pass=total >= reg.course.pass_marks,
                uploaded_by=trainee
            )

        return redirect(
            f"{request.path}?class={selected_class or ''}&section={selected_section or ''}"
        )

    # ✅ PREPARE FINAL DATA
    student_data = []

    for student in students:

        student_regs = eligible_regs.filter(student=student)

        if student_regs.exists():
            for reg in student_regs:

                result = result_map.get((student.id, reg.course.id))

                student_data.append({
                    "student": student,
                    "course": reg.course,
                    "reg_id": reg.id,
                    "marks": result.theory_marks if result else "",
                    "is_locked": True if result else False,
                    "is_detained": False
                })

        else:
            # ❌ DETAINED STUDENT
            student_data.append({
                "student": student,
                "course": None,
                "reg_id": None,
                "marks": "",
                "is_locked": True,
                "is_detained": True
            })

    # ✅ CLASSES
    classes = (
        base_students
        .values_list('student_class', flat=True)
        .distinct()
        .order_by('student_class')
    )

    # ✅ SECTIONS (FILTER BY CLASS)
    section_queryset = base_students

    if selected_class:
        section_queryset = section_queryset.filter(student_class=selected_class)

    sections = (
        section_queryset
        .values_list('section', flat=True)
        .distinct()
        .order_by('section')
    )

    return render(request, "results/trainee_results.html", {
        "student_data": student_data,
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section,
    })
    
    
from django.shortcuts import render
from apps.accounts.models import User
from apps.results.models import Result
from apps.exams.models import ExamRegistration
from apps.results.utils import calculate_grade

from django.shortcuts import render
from apps.accounts.models import User
from apps.results.models import Result
from apps.exams.models import ExamRegistration
from apps.courses.models import CourseRegistration
from apps.results.utils import calculate_grade
from django.shortcuts import render
from apps.accounts.models import User
from apps.results.models import Result, ResultPublish
from apps.courses.models import CourseRegistration
from apps.results.utils import calculate_grade


def school_results(request):

    school = request.user.school
    selected_class = request.GET.get('class')

    # =========================
    # 🔒 CHECK PUBLISH STATUS
    # =========================
    is_published = ResultPublish.objects.filter(
        school=school,
        is_published=True
    ).exists()

    if not is_published:
        return render(request, "results/school_results.html", {
            "not_published": True,
            "result_data": [],
            "classes": [],
            "selected_class": selected_class
        })

    # =========================
    # STUDENTS (FOR FILTER ONLY)
    # =========================
    students = User.objects.filter(
        role='student',
        school=school
    )

    if selected_class:
        students = students.filter(student_class=selected_class)

    # =========================
    # INTERNAL MARKS (SOURCE)
    # =========================
    registrations = CourseRegistration.objects.filter(
        school=school,
        status='approved'
    ).select_related('student', 'course')

    if selected_class:
        registrations = registrations.filter(student__student_class=selected_class)

    # =========================
    # THEORY MARKS (ONLY LOCKED / FINAL)
    # =========================
    results = Result.objects.filter(
        school=school,
        is_locked=True
    )

    result_map = {
        (r.student_id, r.course_id): r
        for r in results
    }

    # =========================
    # FINAL DATA
    # =========================
    result_data = []

    for reg in registrations:

        student = reg.student
        course = reg.course

        result = result_map.get((student.id, course.id))

        theory = result.theory_marks if result else 0
        internal = reg.internal_obtained or 0

        obtained = theory + internal

        grade, gp = calculate_grade(obtained)

        # ✅ PASS / FAIL
        is_pass = obtained >= course.pass_marks

        result_data.append({
            "student": student,
            "course": course,
            "full_marks": course.full_marks,
            "pass_marks": course.pass_marks,
            "theory": theory,
            "internal": internal,
            "obtained": obtained,
            "grade": grade,
            "gp": gp,
            "is_pass": is_pass
        })

    # =========================
    # CLASS FILTER LIST
    # =========================
    classes = (
        User.objects
        .filter(role='student', school=school)
        .exclude(student_class__isnull=True)
        .exclude(student_class__exact="")
        .values_list('student_class', flat=True)
        .distinct()
        .order_by('student_class')
    )

    return render(request, "results/school_results.html", {
        "result_data": result_data,
        "classes": classes,
        "selected_class": selected_class,
        "not_published": False
    })
    
# apps/results/views.py

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import os

from apps.courses.models import CourseRegistration
from apps.results.models import Result
from apps.results.utils import calculate_grade


def export_school_results_excel(request):

    school = request.user.school
    selected_class = request.GET.get('class')

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # =========================
    # HEADER
    # =========================

    ws.merge_cells('A1:I1')
    ws['A1'] = "Nexoalbs Academy, Gaighat, Udayapur"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:I2')
    ws['A2'] = school.name
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:I3')
    ws['A3'] = getattr(school, 'address', 'School Address')
    ws['A3'].alignment = Alignment(horizontal="center")

    if selected_class:
        ws.merge_cells('A4:I4')
        ws['A4'] = f"Class: {selected_class}"
        ws['A4'].alignment = Alignment(horizontal="center")

    # =========================
    # LOGO
    # =========================
    try:
        logo_path = os.path.join("media", "images", "logo.png")
        img = ExcelImage(logo_path)
        img.height = 60
        img.width = 60
        ws.add_image(img, 'A1')
    except:
        pass

    # =========================
    # TABLE HEADER
    # =========================

    headers = [
        "Student", "Course", "Full", "Pass",
        "Theory", "Internal", "Obtained", "Grade", "CGPA"
    ]

    row_num = 6

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # =========================
    # DATA
    # =========================

    registrations = CourseRegistration.objects.filter(
        school=school,
        status='approved'
    ).select_related('student', 'course')

    results = Result.objects.filter(school=school)

    result_map = {
        (r.student_id, r.course_id): r
        for r in results
    }

    row_num += 1

    for reg in registrations:

        student = reg.student
        course = reg.course

        if selected_class and student.student_class != selected_class:
            continue

        result = result_map.get((student.id, course.id))

        theory = result.theory_marks if result else 0
        internal = reg.internal_obtained
        obtained = theory + internal

        grade, gp = calculate_grade(obtained)

        data = [
            student.first_name,
            course.name,
            course.full_marks,
            course.pass_marks,
            theory,
            internal,
            obtained,
            grade,
            gp
        ]

        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

        row_num += 1

    # =========================
    # RESPONSE
    # =========================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"results_class_{selected_class or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)

    return response

from django.shortcuts import render
from apps.results.models import Result, ResultPublish
from apps.courses.models import CourseRegistration
from apps.results.utils import calculate_grade


def student_results(request):

    student = request.user
    school = student.school

    # =========================
    # 🔒 CHECK PUBLISH STATUS
    # =========================
    is_published = ResultPublish.objects.filter(
        school=school,
        is_published=True
    ).exists()

    if not is_published:
        return render(request, "results/student_results.html", {
            "not_published": True,
            "student": student
        })

    # =========================
    # FINAL RESULTS (LOCKED ONLY)
    # =========================
    results = Result.objects.filter(
        student=student,
        school=school,
        is_locked=True
    ).select_related('course')

    # SAFETY: no results even after publish
    if not results.exists():
        return render(request, "results/student_results.html", {
            "not_published": True,
            "student": student
        })

    # =========================
    # INTERNAL MARKS
    # =========================
    registrations = CourseRegistration.objects.filter(
        student=student,
        school=school,
        status='approved'
    ).select_related('course')

    reg_map = {
        r.course_id: r for r in registrations
    }

    # =========================
    # BUILD RESULT DATA
    # =========================
    result_data = []

    total_gp = 0
    total_subjects = 0
    overall_pass = True

    for r in results:

        course = r.course
        reg = reg_map.get(course.id)

        theory = r.theory_marks or 0
        internal = reg.internal_obtained if reg else 0

        obtained = theory + internal

        grade, gp = calculate_grade(obtained)

        # ✅ PASS / FAIL
        is_pass = obtained >= course.pass_marks

        if not is_pass:
            overall_pass = False

        total_gp += gp
        total_subjects += 1

        result_data.append({
            "course": course,
            "full_marks": course.full_marks,
            "pass_marks": course.pass_marks,
            "theory": theory,
            "internal": internal,
            "obtained": obtained,
            "grade": grade,
            "gp": gp,
            "is_pass": is_pass
        })

    # =========================
    # CGPA
    # =========================
    cgpa = round(total_gp / total_subjects, 2) if total_subjects else 0

    return render(request, "results/student_results.html", {
        "result_data": result_data,
        "cgpa": cgpa,
        "student": student,
        "overall_pass": overall_pass,
        "not_published": False
    })